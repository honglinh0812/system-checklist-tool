from flask import Blueprint, request
from flask_jwt_extended import jwt_required
from sqlalchemy import or_, desc
from datetime import datetime, timezone, timedelta

# GMT+7 timezone
GMT_PLUS_7 = timezone(timedelta(hours=7))
from models.mop import Command, MOP
from models.execution import ExecutionHistory
from models.user import User
from models import db
from .api_utils import (
    api_response, api_error, paginate_query, validate_json,
    get_request_filters, apply_filters, require_role
)
from core.schemas import (
    CommandCreateSchema, CommandUpdateSchema, CommandSchema,
    ExecutionCreateSchema, ExecutionSchema
)
from core.auth import get_current_user
import logging
import subprocess
import threading
import time

logger = logging.getLogger(__name__)

commands_bp = Blueprint('commands', __name__, url_prefix='/api/commands')
executions_bp = Blueprint('executions', __name__, url_prefix='/api/executions')

# Command Management
@commands_bp.route('', methods=['GET'])
@jwt_required()
def get_commands():
    """Get paginated list of commands with filtering"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        # Get filter parameters
        filters = get_request_filters()
        
        # Build base query
        query = Command.query.join(MOP)
        
        # Apply role-based filtering
        if current_user.role == 'user':
            # Users can only see commands from their own MOPs
            query = query.filter(MOP.created_by == current_user.id)
        
        # Apply search filter
        if filters.get('search'):
            search_term = f"%{filters['search']}%"
            query = query.filter(
                or_(
                    Command.command_text.ilike(search_term),
                    Command.description.ilike(search_term),
                    MOP.name.ilike(search_term)
                )
            )
        
        # Apply MOP filter
        mop_id = request.args.get('mop_id', type=int)
        if mop_id:
            query = query.filter(Command.mop_id == mop_id)
        
        # Apply critical filter
        is_critical = request.args.get('is_critical')
        if is_critical is not None:
            query = query.filter(Command.is_critical == (is_critical.lower() == 'true'))
        
        # Apply sorting
        sort_by = filters.get('sort_by', 'order_index')
        sort_order = filters.get('sort_order', 'asc')
        
        if hasattr(Command, sort_by):
            column = getattr(Command, sort_by)
            if sort_order.lower() == 'desc':
                query = query.order_by(column.desc())
            else:
                query = query.order_by(column.asc())
        
        # Paginate
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 20, type=int), 100)
        
        result = paginate_query(query, page, per_page)
        
        # Serialize commands
        command_schema = CommandSchema(many=True)
        commands_data = command_schema.dump(result['items'])
        
        # Add MOP info to each command
        for i, command in enumerate(result['items']):
            commands_data[i]['mop_name'] = command.mop.name
            commands_data[i]['mop_status'] = command.mop.status
        
        return api_response({
            'commands': commands_data,
            'pagination': result['pagination']
        })
        
    except Exception as e:
        logger.error(f"Get commands error: {str(e)}")
        return api_error('Failed to fetch commands', 500)

@commands_bp.route('/<int:command_id>', methods=['GET'])
@jwt_required()
def get_command(command_id):
    """Get command details by ID"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        command = Command.query.get(command_id)
        if not command:
            return api_error('Command not found', 404)
        
        # Check permissions
        if current_user.role == 'user' and command.mop.created_by != current_user.id:
            return api_error('Insufficient permissions', 403)
        
        command_schema = CommandSchema()
        command_data = command_schema.dump(command)
        
        # Add MOP info
        command_data['mop_name'] = command.mop.name
        command_data['mop_status'] = command.mop.status
        
        # Add execution history
        executions = ExecutionHistory.query.filter_by(command_id=command_id).order_by(desc(ExecutionHistory.executed_at)).limit(10).all()
        execution_schema = ExecutionSchema(many=True)
        command_data['recent_executions'] = execution_schema.dump(executions)
        
        return api_response(command_data)
        
    except Exception as e:
        logger.error(f"Get command error: {str(e)}")
        return api_error('Failed to fetch command', 500)

@commands_bp.route('/<int:command_id>/execute', methods=['POST'])
@jwt_required()
def execute_command(command_id):
    """Execute a command"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        command = Command.query.get(command_id)
        if not command:
            return api_error('Command not found', 404)
        
        # Check permissions
        if current_user.role == 'user' and command.mop.created_by != current_user.id:
            return api_error('Insufficient permissions', 403)
        
        # Check if MOP is approved for execution
        if command.mop.status not in ['approved', 'in_progress']:
            return api_error('MOP must be approved before executing commands', 400)
        
        # Get execution parameters
        data = request.get_json() or {}
        dry_run = data.get('dry_run', False)
        server_id = data.get('server_id')
        
        # Create execution record
        execution = ExecutionHistory(
            command_id=command_id,
            mop_id=command.mop_id,
            executed_by=current_user.id,
            server_id=server_id,
            status='running',
            dry_run=dry_run
        )
        
        db.session.add(execution)
        db.session.commit()
        
        # Execute command in background
        def execute_in_background():
            try:
                start_time = time.time()
                
                if dry_run:
                    # Simulate execution for dry run
                    time.sleep(1)
                    execution.status = 'completed'
                    execution.exit_code = 0
                    execution.output = f"[DRY RUN] Would execute: {command.command_text}"
                else:
                    # Execute actual command
                    process = subprocess.run(
                        command.command_text,
                        shell=True,
                        capture_output=True,
                        text=True,
                        timeout=command.timeout_seconds or 300
                    )
                    
                    execution.exit_code = process.returncode
                    execution.output = process.stdout
                    execution.error_output = process.stderr
                    
                    if process.returncode == 0:
                        execution.status = 'completed'
                    else:
                        execution.status = 'failed'
                
                execution.duration = time.time() - start_time
                execution.completed_at = datetime.now(GMT_PLUS_7)
                
                db.session.commit()
                
                logger.info(f"Command executed: {command.command_text} by {current_user.username} - Status: {execution.status}")
                
            except subprocess.TimeoutExpired:
                execution.status = 'timeout'
                execution.error_output = 'Command execution timed out'
                execution.completed_at = datetime.now(GMT_PLUS_7)
                db.session.commit()
                logger.warning(f"Command timeout: {command.command_text}")
                
            except Exception as e:
                execution.status = 'error'
                execution.error_output = str(e)
                execution.completed_at = datetime.now(GMT_PLUS_7)
                db.session.commit()
                logger.error(f"Command execution error: {str(e)}")
        
        # Start background execution
        thread = threading.Thread(target=execute_in_background)
        thread.daemon = True
        thread.start()
        
        execution_schema = ExecutionSchema()
        execution_data = execution_schema.dump(execution)
        
        return api_response(execution_data, 'Command execution started', 202)
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Execute command error: {str(e)}")
        return api_error('Failed to execute command', 500)

@commands_bp.route('/<int:command_id>/validate', methods=['POST'])
@jwt_required()
def validate_command(command_id):
    """Validate command syntax and safety"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        command = Command.query.get(command_id)
        if not command:
            return api_error('Command not found', 404)
        
        # Check permissions
        if current_user.role == 'user' and command.mop.created_by != current_user.id:
            return api_error('Insufficient permissions', 403)
        
        # Basic command validation
        validation_results = {
            'is_valid': True,
            'warnings': [],
            'errors': [],
            'suggestions': []
        }
        
        command_text = command.command_text.strip()
        
        # Check for dangerous commands
        dangerous_commands = ['rm -rf', 'dd if=', 'mkfs', 'fdisk', 'format', '> /dev/']
        for dangerous in dangerous_commands:
            if dangerous in command_text.lower():
                validation_results['warnings'].append(f"Potentially dangerous command detected: {dangerous}")
        
        # Check for missing sudo when needed
        system_commands = ['systemctl', 'service', 'mount', 'umount', 'iptables']
        for sys_cmd in system_commands:
            if command_text.startswith(sys_cmd) and not command_text.startswith('sudo'):
                validation_results['suggestions'].append(f"Consider using 'sudo' with {sys_cmd}")
        
        # Check for command existence (basic)
        first_word = command_text.split()[0] if command_text.split() else ''
        if first_word and not first_word.startswith('/') and first_word not in ['sudo', 'cd', 'echo', 'export']:
            try:
                subprocess.run(['which', first_word], capture_output=True, check=True)
            except subprocess.CalledProcessError:
                validation_results['warnings'].append(f"Command '{first_word}' may not be available")
        
        # Check for empty command
        if not command_text:
            validation_results['is_valid'] = False
            validation_results['errors'].append('Command cannot be empty')
        
        return api_response(validation_results)
        
    except Exception as e:
        logger.error(f"Validate command error: {str(e)}")
        return api_error('Failed to validate command', 500)

# Additional command routes from app.py
@commands_bp.route('/validate', methods=['POST'])
def validate_command_text():
    """Validate a shell command"""
    try:
        from services.command_validator import CommandValidator
        
        data = request.get_json()
        if not data or 'command' not in data:
            return api_error('Command is required', 400)
        
        command = data['command'].strip()
        
        # Validate command
        command_validator = CommandValidator()
        validation_result = command_validator.validate_command(command)
        
        return api_response(validation_result)
        
    except Exception as e:
        logger.error(f"Error validating command: {str(e)}")
        return api_error('Internal server error', 500)

@commands_bp.route('/run', methods=['POST'])
@jwt_required()
def run_commands():
    """Run commands on selected servers"""
    from flask_jwt_extended import get_jwt_identity
    from services.ansible_manager import AnsibleRunner
    from services.command_validator import CommandValidator
    import app  # Import to access global variables
    
    current_user_id = get_jwt_identity()
    
    try:
        data = request.get_json()
        if not data:
            return api_error('No data provided', 400)
        
        selected_servers = data.get('selected_servers', [])
        commands = data.get('commands', [])
        mop_id = data.get('mop_id')  # Optional MOP ID
        
        if not selected_servers:
            return api_error('No servers selected', 400)
        
        if not commands:
            return api_error('No commands provided', 400)
        
        command_validator = CommandValidator()
        
        for cmd in commands:
            if 'command' not in cmd or not cmd['command'].strip():
                return api_error('Tất cả lệnh phải có nội dung câu lệnh', 400)
            
            validation_result = command_validator.validate_command(cmd['command'])
            if not validation_result['valid']:
                error_msg = f'Lệnh không hợp lệ: {cmd.get("title", "Không xác định")}'
                if validation_result.get('syntax_error'):
                    error_msg += f' - Lỗi cú pháp: {validation_result["syntax_error"]}'
                elif validation_result.get('errors'):
                    error_msg += f' - {", ".join(validation_result["errors"])}'
                
                return api_error(error_msg, 400, validation_result)
        
        # Get current servers from global storage
        servers_to_run = []
        for server in app.current_servers:
            if server['ip'] in selected_servers:
                servers_to_run.append(server)
        
        if not servers_to_run:
            return api_error('No valid servers found', 400)
        
        # Update global commands
        app.current_commands = commands
        
        timestamp = datetime.now().strftime("%H%M%S_%d%m%Y")
        job_id = f"job_{timestamp}"
        
        # Create execution history record
        from models.execution import ExecutionHistory
        execution = ExecutionHistory(
            mop_id=mop_id,
            user_id=current_user_id,
            risk_assessment=data.get('risk_assessment', False),
            handover_assessment=data.get('handover_assessment', False)
        )
        db.session.add(execution)
        db.session.commit()  # Get the ID
        
        # Run commands in background
        ansible_runner = AnsibleRunner()
        thread = threading.Thread(
            target=ansible_runner.run_playbook,
            args=(job_id, commands, servers_to_run, timestamp, execution.id)
        )
        thread.daemon = True
        thread.start()
        
        return api_response({
            'job_id': job_id,
            'execution_id': execution.id,
            'servers_count': len(servers_to_run),
            'commands_count': len(commands)
        }, 'Commands started successfully')
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error running commands: {str(e)}")
        return api_error('Internal server error', 500)

@commands_bp.route('/status/<job_id>', methods=['GET'])
def get_command_status(job_id):
    """Get status of command execution"""
    try:
        from services.ansible_manager import AnsibleRunner
        
        ansible_runner = AnsibleRunner()
        status = ansible_runner.get_job_status(job_id)
        if status:
            return api_response({
                'status': status
            })
        else:
            return api_error('Job not found', 404)
            
    except Exception as e:
        logger.error(f"Error getting job status: {str(e)}")
        return api_error('Internal server error', 500)

@commands_bp.route('/results/<job_id>', methods=['GET'])
def get_command_results(job_id):
    """Get results of command execution"""
    try:
        from services.ansible_manager import AnsibleRunner
        
        ansible_runner = AnsibleRunner()
        results = ansible_runner.get_job_results(job_id)
        if results:
            return api_response({
                'results': results
            })
        else:
            return api_error('Results not found', 404)
            
    except Exception as e:
        logger.error(f"Error getting job results: {str(e)}")
        return api_error('Internal server error', 500)

# Note: The logs route is kept in app.py as /api/logs/<job_id> to maintain compatibility
# This is because it's used by other parts of the system beyond just commands

# Execution History Management
@executions_bp.route('', methods=['GET'])
@jwt_required()
def get_executions():
    """Get paginated list of execution history"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        # Get filter parameters
        filters = get_request_filters()
        
        # Build base query
        query = ExecutionHistory.query.join(MOP, ExecutionHistory.mop_id == MOP.id)
        
        # Apply role-based filtering
        if current_user.role == 'user':
            # Users can only see executions from their own MOPs
            query = query.filter(MOP.created_by == current_user.id)
        
        # Apply search filter
        if filters.get('search'):
            search_term = f"%{filters['search']}%"
            query = query.filter(
                MOP.name.ilike(search_term)
            )
        
        # Apply status filter
        if filters.get('status'):
            query = query.filter(ExecutionHistory.status == filters['status'])
        
        # Apply MOP filter
        mop_id = request.args.get('mop_id', type=int)
        if mop_id:
            query = query.filter(ExecutionHistory.mop_id == mop_id)
        
        # Apply user filter
        user_id = request.args.get('user_id', type=int)
        if user_id:
            query = query.filter(ExecutionHistory.user_id == user_id)
        
        # Apply date range filter
        if filters.get('date_from'):
            query = query.filter(ExecutionHistory.execution_time >= filters['date_from'])
        if filters.get('date_to'):
            query = query.filter(ExecutionHistory.execution_time <= filters['date_to'])
        
        # Apply sorting
        sort_by = filters.get('sort_by', 'execution_time')
        sort_order = filters.get('sort_order', 'desc')
        
        if hasattr(ExecutionHistory, sort_by):
            column = getattr(ExecutionHistory, sort_by)
            if sort_order.lower() == 'desc':
                query = query.order_by(column.desc())
            else:
                query = query.order_by(column.asc())
        
        # Paginate
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 20, type=int), 100)
        
        result = paginate_query(query, page, per_page)
        
        # Serialize executions
        execution_schema = ExecutionSchema(many=True)
        executions_data = execution_schema.dump(result['items'])
        
        # Add related info to each execution
        for i, execution in enumerate(result['items']):
            executions_data[i]['mop_name'] = execution.mop.name
            # Get user info
            user = User.query.get(execution.user_id)
            executions_data[i]['executor_username'] = user.username if user else 'Unknown'
        
        return api_response({
            'executions': executions_data,
            'pagination': result['pagination']
        })
        
    except Exception as e:
        logger.error(f"Get executions error: {str(e)}")
        return api_error('Failed to fetch executions', 500)

@executions_bp.route('/<int:execution_id>', methods=['GET'])
@jwt_required()
def get_execution(execution_id):
    """Get execution details by ID"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        execution = ExecutionHistory.query.get(execution_id)
        if not execution:
            return api_error('Execution not found', 404)
        
        # Check permissions
        if current_user.role == 'user' and execution.mop.created_by != current_user.id:
            return api_error('Insufficient permissions', 403)
        
        execution_schema = ExecutionSchema()
        execution_data = execution_schema.dump(execution)
        
        # Add related info
        execution_data['command_text'] = execution.command.command_text
        execution_data['command_description'] = execution.command.description
        execution_data['mop_name'] = execution.mop.name
        execution_data['executor_username'] = execution.executed_by_user.username if execution.executed_by_user else 'Unknown'
        
        return api_response(execution_data)
        
    except Exception as e:
        logger.error(f"Get execution error: {str(e)}")
        return api_error('Failed to fetch execution', 500)

@executions_bp.route('/stats', methods=['GET'])
@jwt_required()
def get_execution_stats():
    """Get execution statistics"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        # Build base query
        query = ExecutionHistory.query.join(Command).join(MOP)
        
        # Apply role-based filtering
        if current_user.role == 'user':
            query = query.filter(MOP.created_by == current_user.id)
        
        # Get date range
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        if date_from:
            query = query.filter(ExecutionHistory.executed_at >= date_from)
        if date_to:
            query = query.filter(ExecutionHistory.executed_at <= date_to)
        
        # Calculate statistics
        total_executions = query.count()
        successful_executions = query.filter(ExecutionHistory.status == 'completed').count()
        failed_executions = query.filter(ExecutionHistory.status == 'failed').count()
        running_executions = query.filter(ExecutionHistory.status == 'running').count()
        
        # Calculate success rate
        success_rate = (successful_executions / total_executions * 100) if total_executions > 0 else 0
        
        # Get average execution time
        completed_executions = query.filter(
            ExecutionHistory.status == 'completed',
            ExecutionHistory.duration.isnot(None)
        ).all()
        
        avg_duration = 0
        if completed_executions:
            total_duration = sum(ex.duration for ex in completed_executions)
            avg_duration = total_duration / len(completed_executions)
        
        # Get top failed commands
        failed_commands = db.session.query(
            Command.command_text,
            db.func.count(ExecutionHistory.id).label('failure_count')
        ).join(ExecutionHistory).filter(
            ExecutionHistory.status == 'failed'
        ).group_by(Command.command_text).order_by(
            db.func.count(ExecutionHistory.id).desc()
        ).limit(5).all()
        
        stats = {
            'total_executions': total_executions,
            'successful_executions': successful_executions,
            'failed_executions': failed_executions,
            'running_executions': running_executions,
            'success_rate': round(success_rate, 2),
            'average_duration': round(avg_duration, 2),
            'top_failed_commands': [
                {'command': cmd[0][:100], 'failure_count': cmd[1]}
                for cmd in failed_commands
            ]
        }
        
        return api_response(stats)
        
    except Exception as e:
        logger.error(f"Get execution stats error: {str(e)}")
        return api_error('Failed to fetch execution statistics', 500)

@executions_bp.route('/<int:execution_id>/stop', methods=['POST'])
@jwt_required()
def stop_execution(execution_id):
    """Stop a running execution"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        execution = ExecutionHistory.query.get_or_404(execution_id)
        
        # Check permissions
        if current_user.role == 'user' and execution.executed_by != current_user.id:
            return api_error('Access denied', 403)
        
        if execution.status != 'running':
            return api_error('Execution is not running', 400)
        
        # Update execution status
        execution.status = 'cancelled'
        execution.ended_at = datetime.now(GMT_PLUS_7)
        execution.cancelled_by = current_user.id
        
        # TODO: Implement actual process termination logic here
        # This would involve killing the actual subprocess/ansible process
        
        db.session.commit()
        
        logger.info(f"Execution {execution_id} stopped by user {current_user.id}")
        
        return api_response({
            'message': 'Execution stopped successfully',
            'execution_id': execution_id,
            'status': execution.status
        })
        
    except Exception as e:
        logger.error(f"Error stopping execution {execution_id}: {str(e)}")
        db.session.rollback()
        return api_error('Failed to stop execution', 500)

@executions_bp.route('/<int:execution_id>/logs', methods=['GET'])
@jwt_required()
def get_execution_logs(execution_id):
    """Get real-time logs for an execution"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        execution = ExecutionHistory.query.get_or_404(execution_id)
        
        # Check permissions
        if current_user.role == 'user' and execution.executed_by != current_user.id:
            return api_error('Access denied', 403)
        
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 100, type=int), 500)
        
        # Get log level filter
        log_level = request.args.get('level', 'all')
        
        # TODO: Implement actual log retrieval from log files or database
        # For now, return mock logs based on execution status
        logs = []
        
        if execution.status == 'running':
            logs = [
                {'timestamp': datetime.now(GMT_PLUS_7).isoformat(), 'level': 'INFO', 'message': 'Execution in progress...'},
                {'timestamp': datetime.now(GMT_PLUS_7).isoformat(), 'level': 'DEBUG', 'message': 'Processing commands...'}
            ]
        elif execution.status == 'completed':
            logs = [
                {'timestamp': execution.started_at.isoformat(), 'level': 'INFO', 'message': 'Execution started'},
                {'timestamp': execution.ended_at.isoformat(), 'level': 'INFO', 'message': 'Execution completed successfully'}
            ]
        elif execution.status == 'failed':
            logs = [
                {'timestamp': execution.started_at.isoformat(), 'level': 'INFO', 'message': 'Execution started'},
                {'timestamp': execution.ended_at.isoformat(), 'level': 'ERROR', 'message': 'Execution failed'}
            ]
        
        # Filter by log level if specified
        if log_level != 'all':
            logs = [log for log in logs if log['level'].lower() == log_level.lower()]
        
        # Simple pagination
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_logs = logs[start_idx:end_idx]
        
        return api_response({
            'logs': paginated_logs,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': len(logs),
                'pages': (len(logs) + per_page - 1) // per_page
            },
            'execution_status': execution.status
        })
        
    except Exception as e:
        logger.error(f"Error fetching logs for execution {execution_id}: {str(e)}")
        return api_error('Failed to fetch logs', 500)

# Additional execution routes from app.py
@executions_bp.route('/history', methods=['GET'])
@jwt_required()
def get_execution_history():
    """Get execution history for the last 7 days"""
    try:
        from datetime import timedelta
        from models.execution import ExecutionHistory
        from models.assessment import AssessmentResult
        
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        # Get executions from last 7 days
        seven_days_ago = datetime.now(GMT_PLUS_7) - timedelta(days=7)
        
        # Get ExecutionHistory records
        execution_query = ExecutionHistory.query.filter(
            ExecutionHistory.started_at >= seven_days_ago
        )
        
        # Get AssessmentResult records
        assessment_query = AssessmentResult.query.filter(
            AssessmentResult.created_at >= seven_days_ago
        )
        
        # Apply role-based filtering
        if current_user.role == 'user':
            execution_query = execution_query.filter(ExecutionHistory.executed_by == current_user.id)
            assessment_query = assessment_query.filter(AssessmentResult.executed_by == current_user.id)
        
        executions = execution_query.order_by(ExecutionHistory.started_at.desc()).all()
        assessments = assessment_query.order_by(AssessmentResult.created_at.desc()).all()
        
        execution_list = []
        
        # Process ExecutionHistory records
        for exec in executions:
            # Calculate results summary
            total_commands = exec.total_commands or len(exec.results) if exec.results else 0
            passed_commands = exec.completed_commands or (sum(1 for r in exec.results if r.is_valid) if exec.results else 0)
            failed_commands = total_commands - passed_commands
            success_rate = (passed_commands / total_commands * 100) if total_commands > 0 else 0
            
            # Determine assessment type
            assessment_type = "Đánh giá rủi ro" if exec.risk_assessment else "Đánh giá bàn giao"
            
            # Get user info
            user_name = exec.executed_by_user.username if exec.executed_by_user else 'Unknown'
            
            execution_data = {
                'id': exec.id,
                'type': 'execution',
                'mop_id': exec.mop_id,
                'user_id': exec.executed_by,
                'user_name': user_name,
                'execution_time': exec.started_at.isoformat() if exec.started_at else exec.created_at.isoformat() if hasattr(exec, 'created_at') else None,
                'execution_time_formatted': exec.started_at.strftime('%Y-%m-%d %H:%M:%S') if exec.started_at else 'N/A',
                'risk_assessment': exec.risk_assessment,
                'handover_assessment': exec.handover_assessment,
                'assessment_type': assessment_type,
                'server_count': len(exec.target_servers.split(',')) if exec.target_servers else 0,
                'total_commands': total_commands,
                'passed_commands': passed_commands,
                'failed_commands': failed_commands,
                'success_rate': success_rate,
                'mop_name': exec.mop.name if exec.mop else 'Unknown MOP',
                'status': exec.status.title() if exec.status else 'Unknown',
                'duration': exec.duration
            }
            execution_list.append(execution_data)
        
        # Process AssessmentResult records
        for assessment in assessments:
            # Determine assessment type
            assessment_type = "Đánh giá rủi ro" if assessment.assessment_type == 'risk' else "Đánh giá bàn giao"
            
            # Get user info
            user_name = assessment.executor.username if assessment.executor else 'Unknown'
            
            # Calculate server count
            server_count = len(assessment.server_info) if assessment.server_info else 0
            
            # Calculate success rate from test results
            success_rate = 0
            total_tests = 0
            passed_tests = 0
            
            if assessment.test_results:
                # Handle both dict and list formats for test_results
                if isinstance(assessment.test_results, dict):
                    for server_ip, results in assessment.test_results.items():
                        if isinstance(results, dict):
                            for cmd_result in results.values():
                                if isinstance(cmd_result, dict) and 'status' in cmd_result:
                                    total_tests += 1
                                    if cmd_result['status'] == 'OK':
                                        passed_tests += 1
                elif isinstance(assessment.test_results, list):
                    for result in assessment.test_results:
                        if isinstance(result, dict) and 'status' in result:
                            total_tests += 1
                            if result['status'] == 'OK':
                                passed_tests += 1
                
                success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
            
            execution_data = {
                'id': assessment.id,
                'type': 'assessment',
                'mop_id': assessment.mop_id,
                'user_id': assessment.executed_by,
                'user_name': user_name,
                'execution_time': assessment.started_at.isoformat() if assessment.started_at else assessment.created_at.isoformat(),
                'execution_time_formatted': (assessment.started_at or assessment.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                'risk_assessment': assessment.assessment_type == 'risk',
                'handover_assessment': assessment.assessment_type == 'handover',
                'assessment_type': assessment_type,
                'server_count': server_count,
                'total_commands': total_tests,
                'passed_commands': passed_tests,
                'failed_commands': total_tests - passed_tests,
                'success_rate': success_rate,
                'mop_name': assessment.mop.name if assessment.mop else 'Unknown MOP',
                'status': assessment.status.title() if assessment.status else 'Unknown',
                'duration': (assessment.completed_at - assessment.started_at).total_seconds() if assessment.completed_at and assessment.started_at else None
            }
            execution_list.append(execution_data)
        
        # Sort by execution time (newest first)
        execution_list.sort(key=lambda x: x['execution_time'] or '', reverse=True)
        
        return api_response({
            'executions': execution_list
        })
        
    except Exception as e:
        logger.error(f"Error getting execution history: {str(e)}")
        return api_error('Internal server error', 500)

@executions_bp.route('/<int:execution_id>/detail', methods=['GET'])
@jwt_required()
def get_execution_detail(execution_id):
    """Get detailed execution information"""
    try:
        from models.execution import ExecutionHistory
        
        execution = ExecutionHistory.query.get(execution_id)
        if not execution:
            return api_error('Execution not found', 404)
        
        # Group results by server
        server_results = {}
        for result in execution.results:
            if result.server_ip not in server_results:
                server_results[result.server_ip] = []
            server_results[result.server_ip].append({
                'id': result.id,
                'command_id': result.command_id,
                'output': result.output,
                'stderr': result.stderr,
                'return_code': result.return_code,
                'is_valid': result.is_valid,
                'command': result.command
            })
        
        return api_response({
            'execution': {
                'id': execution.id,
                'mop_id': execution.mop_id,
                'user_id': execution.user_id,
                'execution_time': execution.execution_time.isoformat(),
                'risk_assessment': execution.risk_assessment,
                'handover_assessment': execution.handover_assessment,
                'mop': {
                    'id': execution.mop.id,
                    'name': execution.mop.name,
                    'commands': [{
                        'id': cmd.id,
                        'title': cmd.title,
                        'command': cmd.command,
                        'reference_value': cmd.reference_value
                    } for cmd in execution.mop.commands]
                },
                'user': {
                    'id': execution.user.id,
                    'username': execution.user.username
                },
                'results': [{
                    'server_ip': server_ip,
                    'commands': results
                } for server_ip, results in server_results.items()]
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting execution detail: {str(e)}")
        return api_error('Internal server error', 500)

# MOP Execution endpoint
@executions_bp.route('/mop/<int:mop_id>/execute', methods=['POST'])
@jwt_required()
def execute_mop(mop_id):
    """Execute a MOP with all its commands"""
    try:
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        # Get MOP
        mop = MOP.query.get(mop_id)
        if not mop:
            return api_error('MOP not found', 404)
        
        # Check permissions
        if current_user.role == 'user' and mop.created_by != current_user.id:
            return api_error('Insufficient permissions', 403)
        
        # Get request data
        data = request.get_json() or {}
        servers = data.get('servers', [])
        execution_mode = data.get('execution_mode', 'sequential')
        dry_run = data.get('dry_run', False)
        
        if not servers:
            return api_error('No servers specified', 400)
        
        # Create execution record
        execution = ExecutionHistory(
            mop_id=mop_id,
            executed_by=current_user.id,
            user_id=current_user.id,  # Legacy field for backward compatibility
            risk_assessment=data.get('risk_assessment', False),
            handover_assessment=data.get('handover_assessment', False),
            status='pending',
            target_servers=','.join(servers),
            execution_mode=execution_mode,
            dry_run=dry_run
        )
        db.session.add(execution)
        db.session.flush()
        
        # Start execution in background
        import threading
        from services.ansible_manager import AnsibleRunner
        
        ansible_runner = AnsibleRunner()
        
        def run_execution():
            try:
                # Update status to running
                execution.status = 'running'
                execution.started_at = datetime.now(GMT_PLUS_7)
                db.session.commit()
                
                # Convert MOP commands to execution format
                commands = [{
                    'id': cmd.id,
                    'title': f'Command {cmd.order_index}',
                    'command': cmd.command_text,
                    'description': cmd.description,
                    'is_critical': cmd.is_critical,
                    'timeout': cmd.timeout_seconds or 300
                } for cmd in mop.commands.order_by(Command.order_index)]
                
                execution.total_commands = len(commands)
                db.session.commit()
                
                # Generate job ID
                timestamp = datetime.now().strftime("%H%M%S_%d%m%Y")
                job_id = f"mop_{mop_id}_{timestamp}"
                
                # Run the execution
                ansible_runner.run_playbook(
                    job_id=job_id,
                    commands=commands,
                    servers=servers,
                    timestamp=timestamp,
                    execution_id=execution.id,
                    dry_run=dry_run
                )
                
                # Update status to completed
                execution.status = 'completed'
                execution.completed_at = datetime.now(GMT_PLUS_7)
                if execution.started_at:
                    execution.duration = (execution.completed_at - execution.started_at).total_seconds()
                db.session.commit()
                
            except Exception as e:
                logger.error(f"Error in MOP execution thread: {str(e)}")
                # Update execution status to failed
                execution.status = 'failed'
                execution.completed_at = datetime.now(GMT_PLUS_7)
                if execution.started_at:
                    execution.duration = (execution.completed_at - execution.started_at).total_seconds()
                db.session.commit()
        
        thread = threading.Thread(target=run_execution)
        thread.daemon = True
        thread.start()
        
        return api_response({
            'execution_id': execution.id,
            'mop_id': mop_id,
            'status': 'started',
            'servers_count': len(servers),
            'commands_count': len(mop.commands.all()),
            'dry_run': dry_run
        }, 'MOP execution started successfully')
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error executing MOP: {str(e)}")
        return api_error('Failed to start execution', 500)

@executions_bp.route('/export', methods=['GET'])
@jwt_required()
def export_all_executions():
    """Export all executions to Excel"""
    try:
        from services.excel_exporter import ExcelExporter
        from models.execution import ExecutionHistory
        from models.assessment import AssessmentResult
        from datetime import timedelta
        
        current_user = get_current_user()
        if not current_user:
            return api_error('User not found', 404)
        
        # Get date range from query params (default to last 30 days)
        days = request.args.get('days', 30, type=int)
        start_date = datetime.now(GMT_PLUS_7) - timedelta(days=days)
        
        # Get ExecutionHistory records
        execution_query = ExecutionHistory.query.filter(
            ExecutionHistory.started_at >= start_date
        )
        
        # Get AssessmentResult records
        assessment_query = AssessmentResult.query.filter(
            AssessmentResult.created_at >= start_date
        )
        
        # Apply role-based filtering
        if current_user.role == 'user':
            execution_query = execution_query.filter(ExecutionHistory.executed_by == current_user.id)
            assessment_query = assessment_query.filter(AssessmentResult.executed_by == current_user.id)
        
        executions = execution_query.order_by(ExecutionHistory.started_at.desc()).all()
        assessments = assessment_query.order_by(AssessmentResult.created_at.desc()).all()
        
        # Prepare data for export
        export_data = []
        
        # Process ExecutionHistory records
        for exec in executions:
            total_commands = exec.total_commands or len(exec.results) if exec.results else 0
            passed_commands = exec.completed_commands or (sum(1 for r in exec.results if r.is_valid) if exec.results else 0)
            failed_commands = total_commands - passed_commands
            success_rate = (passed_commands / total_commands * 100) if total_commands > 0 else 0
            
            assessment_type = "Đánh giá rủi ro" if exec.risk_assessment else "Đánh giá bàn giao"
            user_name = exec.executed_by_user.username if exec.executed_by_user else 'Unknown'
            
            export_data.append({
                'ID': exec.id,
                'Loại': assessment_type,
                'Tên MOP': exec.mop.name if exec.mop else 'Unknown',
                'Người thực hiện': user_name,
                'Thời gian thực hiện': exec.started_at.strftime('%Y-%m-%d %H:%M:%S') if exec.started_at else 'N/A',
                'Số server': len(exec.target_servers.split(',')) if exec.target_servers else 0,
                'Tổng số lệnh': total_commands,
                'Lệnh thành công': passed_commands,
                'Lệnh thất bại': failed_commands,
                'Tỷ lệ thành công (%)': round(success_rate, 2),
                'Trạng thái': exec.status.title() if exec.status else 'Unknown',
                'Thời gian hoàn thành': exec.completed_at.strftime('%Y-%m-%d %H:%M:%S') if exec.completed_at else 'N/A'
            })
        
        # Process AssessmentResult records
        for assessment in assessments:
            # Calculate test statistics from test_results
            total_tests = 0
            passed_tests = 0
            
            if assessment.test_results:
                if isinstance(assessment.test_results, list):
                    total_tests = len(assessment.test_results)
                    passed_tests = sum(1 for result in assessment.test_results if result.get('result') == 'success')
                elif isinstance(assessment.test_results, dict):
                    for server_ip, results in assessment.test_results.items():
                        if isinstance(results, dict):
                            for cmd_result in results.values():
                                if isinstance(cmd_result, dict) and 'status' in cmd_result:
                                    total_tests += 1
                                    if cmd_result['status'] == 'OK':
                                        passed_tests += 1
            
            success_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
            
            assessment_type = "Đánh giá rủi ro" if assessment.assessment_type == 'risk' else "Đánh giá bàn giao"
            user_name = assessment.executor.username if assessment.executor else 'Unknown'
            
            # Calculate server count from server_info
            server_count = len(assessment.server_info) if assessment.server_info else 0
            
            export_data.append({
                'ID': f"A{assessment.id}",
                'Loại': assessment_type,
                'Tên MOP': assessment.mop.name if assessment.mop else 'Unknown',
                'Người thực hiện': user_name,
                'Thời gian thực hiện': assessment.created_at.strftime('%Y-%m-%d %H:%M:%S') if assessment.created_at else 'N/A',
                'Số server': server_count,
                'Tổng số lệnh': total_tests,
                'Lệnh thành công': passed_tests,
                'Lệnh thất bại': total_tests - passed_tests,
                'Tỷ lệ thành công (%)': round(success_rate, 2),
                'Trạng thái': assessment.status.title() if assessment.status else 'Unknown',
                'Thời gian hoàn thành': assessment.completed_at.strftime('%Y-%m-%d %H:%M:%S') if assessment.completed_at else 'N/A'
            })
        
        # Create Excel file
        excel_exporter = ExcelExporter()
        timestamp = datetime.now(GMT_PLUS_7).strftime('%Y%m%d_%H%M%S')
        filename = f"execution_history_{timestamp}.xlsx"
        
        # Use a simple export method for tabular data
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import send_file
        import os
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lịch sử thực hiện"
        
        # Headers
        headers = list(export_data[0].keys()) if export_data else []
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, data in enumerate(export_data, 2):
            for col, value in enumerate(data.values(), 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        export_dir = os.path.join(os.getcwd(), 'config', 'reports')
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"Error exporting executions: {str(e)}")
        return api_error('Internal server error', 500)