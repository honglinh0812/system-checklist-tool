import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timezone, timedelta
import os
import logging
from typing import Dict, List, Any, Optional

# GMT+7 timezone
GMT_PLUS_7 = timezone(timedelta(hours=7))

logger = logging.getLogger(__name__)

class ExcelExporter:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        # Locale for localized text in exports
        self.locale = 'en'

    def set_locale(self, locale_code: str):
        try:
            if locale_code:
                self.locale = locale_code.split('-')[0].lower()
        except Exception:
            self.locale = 'en'

    def _format_reference_value(self, value: Any, comparator_method: str = '') -> str:
        text = '' if value is None else str(value)
        comp = (comparator_method or '').strip().lower()
        if comp == 'empty':
            return 'Rỗng' if self.locale == 'vi' else 'Empty'
        if comp in ('non_empty', 'not_empty'):
            return 'Không rỗng' if self.locale == 'vi' else 'Not empty'
        # fallback: show raw or Empty when truly blank
        is_empty = len(text.strip()) == 0
        if self.locale == 'vi':
            return 'Rỗng' if is_empty else text
        return 'Empty' if is_empty else text
    
    def export_execution_results(self, execution_data: Dict[str, Any], filename: str = None) -> str:
        """
        Export execution results to Excel with detailed formatting
        
        Args:
            execution_data: Dictionary containing execution results
            filename: Optional filename for the export
            
        Returns:
            Path to the exported Excel file
        """
        try:
            if not filename:
                timestamp = datetime.now(GMT_PLUS_7).strftime("%Y%m%d_%H%M%S")
                filename = f"execution_results_{timestamp}.xlsx"
            
            # Create workbook and worksheets
            wb = openpyxl.Workbook()
            
            # Summary sheet
            self._create_summary_sheet(wb, execution_data)
            
            # Detailed results sheet
            self._create_detailed_sheet(wb, execution_data)
            
            # Server summary sheet
            self._create_server_summary_sheet(wb, execution_data)
            
            # Matrix sheet per requirement (commands x IP, OK/Not OK + final row)
            self._create_matrix_sheet(wb, execution_data)
            
            # Save the workbook
            filepath = f"exports/{filename}"
            wb.save(filepath)
            
            logger.info(f"Excel export completed: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create summary sheet with execution overview"""
        ws = wb.active
        ws.title = "Summary"
        
        # Execution information
        row = 1
        ws.cell(row=row, column=1, value="Execution Summary").font = Font(bold=True, size=14)
        row += 2
        
        summary_data = [
            ["MOP Name", data.get('mop_name', 'N/A')],
            ["Execution Type", data.get('execution_type', 'N/A')],
            ["Executed By", data.get('executed_by', 'N/A')],
            ["Execution Time", data.get('execution_time', 'N/A')],
            ["Total Servers", str(data.get('total_servers', 0))],
            ["Total Commands", str(data.get('total_commands', 0))],
            ["OK Commands", str(data.get('passed_commands', 0))],
            ["Not OK Commands", str(data.get('failed_commands', 0))],
            ["Success Rate", f"{data.get('success_rate', 0):.1f}%"],
            ["Average Score", f"{data.get('average_score', 0):.1f}%"]
        ]
        
        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Validation summary
        row += 2
        ws.cell(row=row, column=1, value="Validation Summary").font = Font(bold=True, size=12)
        row += 1
        
        validation_summary = data.get('validation_summary', {})
        if validation_summary:
            headers = ["Validation Type", "Total", "OK", "Not OK", "Success Rate"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            row += 1
            
            for vtype, stats in validation_summary.items():
                total = stats.get('total', 0)
                passed = stats.get('passed', 0)
                failed = total - passed
                success_rate = (passed / total * 100) if total > 0 else 0
                
                ws.cell(row=row, column=1, value=vtype).border = self.border
                ws.cell(row=row, column=2, value=total).border = self.border
                ws.cell(row=row, column=3, value=passed).border = self.border
                ws.cell(row=row, column=4, value=failed).border = self.border
                ws.cell(row=row, column=5, value=f"{success_rate:.1f}%").border = self.border
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_detailed_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create detailed results sheet with all command outputs"""
        ws = wb.create_sheet("Detailed Results")
        
        # Headers
        headers = [
            "Server IP", "Command Title", "Command", "Expected Output", 
            "Actual Output", "Validation Type", "Result", "Decision", "Score", "Details"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Data rows
        row = 2
        results = data.get('results', [])
        
        for result in results:
            # Handle skipped commands
            is_skipped = result.get('skipped', False)
            skip_reason = result.get('skip_reason', '')
            
            if is_skipped:
                result_status = "SKIPPED"
                decision = "OK (skipped)"
                status_color = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow for skipped
                # Create details with skip condition info
                skip_condition = result.get('skip_condition', {})
                condition_id = skip_condition.get('condition_id', '')
                condition_type = skip_condition.get('condition_type', '')
                details = f"Skipped due to condition: {condition_id} result is {condition_type}" if condition_id and condition_type else skip_reason
            else:
                # Status color coding
                result_status = "OK" if result.get('is_valid', False) else "Not OK"
                decision = result.get('decision', 'APPROVED' if result.get('is_valid', False) else 'REJECTED')
                status_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if result_status == "OK" else PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                details = str(result.get('details', ''))
            
            # Row data
            row_data = [
                result.get('server_ip', ''),
                result.get('command_title', ''),
                result.get('command', ''),
                self._format_reference_value(result.get('expected_output', ''), result.get('comparator_method', '')),
                result.get('actual_output', ''),
                result.get('validation_type', ''),
                result_status,
                decision,
                f"{result.get('score', 0):.1f}%",
                details
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col == 7:  # Result column
                    cell.fill = status_color
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_server_summary_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create server summary sheet with per-server statistics"""
        ws = wb.create_sheet("Server Summary")
        
        # Headers
        headers = [
            "Server IP", "Total Commands", "Passed Commands", "Failed Commands", 
            "Success Rate", "Average Score", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Group results by server
        server_stats = {}
        results = data.get('results', [])
        
        for result in results:
            server_ip = result.get('server_ip', 'Unknown')
            if server_ip not in server_stats:
                server_stats[server_ip] = {
                    'total': 0,
                    'passed': 0,
                    'failed': 0,
                    'scores': []
                }
            
            server_stats[server_ip]['total'] += 1
            if result.get('is_valid', False):
                server_stats[server_ip]['passed'] += 1
            else:
                server_stats[server_ip]['failed'] += 1
            
            server_stats[server_ip]['scores'].append(result.get('score', 0))
        
        # Create rows for each server
        row = 2
        for server_ip, stats in server_stats.items():
            success_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
            avg_score = sum(stats['scores']) / len(stats['scores']) if stats['scores'] else 0
            status = "PASS" if success_rate == 100 else "PARTIAL" if success_rate > 0 else "FAIL"
            
            # Status color coding
            if status == "PASS":
                status_color = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status == "PARTIAL":
                status_color = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                status_color = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row_data = [
                server_ip,
                stats['total'],
                stats['passed'],
                stats['failed'],
                f"{success_rate:.1f}%",
                f"{avg_score:.1f}%",
                status
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col == 7:  # Status column
                    cell.fill = status_color
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_matrix_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create matrix sheet theo format yêu cầu: 
        - Ô (1,1) trống
        - Cột 1 từ hàng 2: tiêu đề lệnh
        - Cột 2 từ hàng 2: câu lệnh
        - Từ cột 3 trở đi: IP server
        - Hàng cuối: kết quả tổng thể
        """
        ws = wb.create_sheet("Assessment Report")
        results = data.get('results', [])

        # Collect unique server IPs and commands
        server_ips: List[str] = []
        commands_data: List[Dict] = []
        
        # Group results by command to get unique commands
        commands_dict = {}
        for r in results:
            title = r.get('command_title', '')
            command = r.get('command', '')
            if title not in commands_dict:
                commands_dict[title] = {
                    'title': title,
                    'command': command,
                    'results': {}
                }
            
            ip = r.get('server_ip', '')
            if ip and ip not in server_ips:
                server_ips.append(ip)
            
            # Store result for this command-server combination
            is_valid = bool(r.get('is_valid', False))
            commands_dict[title]['results'][ip] = is_valid
        
        commands_data = list(commands_dict.values())

        # Ô (1,1) để trống
        # Hàng 1: headers cho IP servers từ cột 3 trở đi
        for col, ip in enumerate(server_ips, 3):
            cell = ws.cell(row=1, column=col, value=ip)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment

        # Từ hàng 2 trở đi: dữ liệu commands
        row = 2
        for cmd_data in commands_data:
            # Cột 1: tiêu đề lệnh
            cell1 = ws.cell(row=row, column=1, value=cmd_data['title'])
            cell1.border = self.border
            
            # Cột 2: câu lệnh
            cell2 = ws.cell(row=row, column=2, value=cmd_data['command'])
            cell2.border = self.border
            
            # Từ cột 3 trở đi: kết quả cho từng server
            for col, ip in enumerate(server_ips, 3):
                is_ok = cmd_data['results'].get(ip, False)
                cell = ws.cell(row=row, column=col, value="OK" if is_ok else "Not OK")
                cell.border = self.border
                cell.fill = PatternFill(
                    start_color="C6EFCE" if is_ok else "FFC7CE", 
                    end_color="C6EFCE" if is_ok else "FFC7CE", 
                    fill_type="solid"
                )
                cell.alignment = self.center_alignment
            
            row += 1

        # Hàng cuối: "Kết quả kiểm tra cuối"
        cell_final = ws.cell(row=row, column=1, value="Kết quả kiểm tra cuối")
        cell_final.font = Font(bold=True)
        cell_final.border = self.border
        
        # Tính kết quả tổng thể cho từng server
        for col, ip in enumerate(server_ips, 3):
            # Kiểm tra tất cả kết quả của server này
            all_ok = True
            for check_row in range(2, row):
                if ws.cell(row=check_row, column=col).value != "OK":
                    all_ok = False
                    break
            
            cell = ws.cell(row=row, column=col, value="OK" if all_ok else "Not OK")
            cell.border = self.border
            cell.fill = PatternFill(
                start_color="C6EFCE" if all_ok else "FFC7CE", 
                end_color="C6EFCE" if all_ok else "FFC7CE", 
                fill_type="solid"
            )
            cell.alignment = self.center_alignment
            cell.font = Font(bold=True)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for c in column:
                try:
                    if len(str(c.value)) > max_length:
                        max_length = len(str(c.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_mop_template(self, mop_data: Dict[str, Any], filename: str = None) -> str:
        """
        Export MOP template to Excel
        
        Args:
            mop_data: MOP data to export
            filename: Optional filename
            
        Returns:
            Path to the exported Excel file
        """
        try:
            if not filename:
                timestamp = datetime.now(GMT_PLUS_7).strftime("%Y%m%d_%H%M%S")
                filename = f"mop_template_{timestamp}.xlsx"
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "MOP Template"
            
            # MOP Information
            row = 1
            ws.cell(row=row, column=1, value="MOP Information").font = Font(bold=True, size=14)
            row += 2
            
            mop_info = [
                ["Name", mop_data.get('name', '')],
                ["Type", mop_data.get('type', '')],
                ["Description", mop_data.get('description', '')],
                ["Status", mop_data.get('status', '')],
                ["Created By", mop_data.get('created_by', '')],
                ["Created Date", mop_data.get('created_at', '')]
            ]
            
            for label, value in mop_info:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1
            
            # Commands section
            row += 2
            ws.cell(row=row, column=1, value="Commands").font = Font(bold=True, size=12)
            row += 1
            
            # Command headers
            cmd_headers = ["Title", "Command", "Expected Output", "Validation Type"]
            for col, header in enumerate(cmd_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            row += 1
            
            # Command data
            commands = mop_data.get('commands', [])
            for command in commands:
                row_data = [
                    command.get('title', ''),
                    command.get('command', ''),
                    command.get('expected_output', ''),
                    command.get('validation_type', 'exact_match')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                
                row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filepath = f"exports/{filename}"
            wb.save(filepath)
            
            logger.info(f"MOP template export completed: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting MOP template: {str(e)}")
            raise
    
    def export_user_report(self, users_data: List[Dict[str, Any]], filename: str = None) -> str:
        """
        Export user management report to Excel
        
        Args:
            users_data: List of user data
            filename: Optional filename
            
        Returns:
            Path to the exported Excel file
        """
        try:
            if not filename:
                timestamp = datetime.now(GMT_PLUS_7).strftime("%Y%m%d_%H%M%S")
                filename = f"user_report_{timestamp}.xlsx"
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "User Report"
            
            # Headers
            headers = [
                "Username", "Email", "Role", "Status", "Last Login", "Created Date"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            # Data rows
            row = 2
            for user in users_data:
                row_data = [
                    user.get('username', ''),
                    user.get('email', ''),
                    user.get('role', ''),
                    'Active' if user.get('is_active', False) else 'Inactive',
                    user.get('last_login', 'Never'),
                    user.get('created_at', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                
                row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filepath = f"exports/{filename}"
            wb.save(filepath)
            
            logger.info(f"User report export completed: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting user report: {str(e)}")
            raise
